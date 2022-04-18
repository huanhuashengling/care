from typing import TYPE_CHECKING
from django.shortcuts import render
from .models import Sclass, Student, Teacher, EvaluateIndex, EvaluateItem, EvaluateLog
from openpyxl import load_workbook
from django.db import connection
import os
import sys
import jieba
import jieba.analyse as analyse
import jieba.posseg as pseg
from optparse import OptionParser

sys.path.append('../')

def read(request):
    readStudentData()
    readTeacherData()
    return render(request, 'picks/read.html')

def index(request):
    # clear all feedbacks
    EvaluateLog.objects.all().delete()
    
    cursor = connection.cursor()
    cursor.execute("TRUNCATE TABLE `evaluate_logs`")

    wb_list = []
    path = "/Users/ywj/Sites/care/static/care_data" #文件夹目录
    files= os.listdir(path) #得到文件夹下的所有文件名称
    for file in files: #遍历文件夹
        if file.endswith('.xlsx'):
            wb = load_workbook(path+"/"+file)
            sheet_obj = wb.active
            # max_col = sheet_obj.max_column
            max_row = sheet_obj.max_row

            class_title = sheet_obj.cell(row = 1, column = 2).value
            tClass = Sclass.objects.filter(display_title = class_title)[0]
            print(file, class_title)
            for i in range(4, max_row + 1):
                tStudentName = sheet_obj.cell(row = i, column = 1).value
                if "哦哦" == tStudentName:
                    continue

                if "丁博棋" == tStudentName:
                    tStudentName = "丁镈棋"

                if "迪力木热提" == tStudentName or "10迪力木热提" == tStudentName:
                    tStudentName = "迪力木热提·艾则孜"
                # print(i, tStudentName)
                # tStudent = Student.objects.filter(student_fname = tStudentName, sclasses_id = tClass.id).row()
                tStudentObj = Student.objects.raw("""SELECT * FROM students where %s like concat('%%', student_name, '%%') and sclasses_id=%s""", [tStudentName, tClass.id])
                if not tStudentObj:
                    print(i, tStudentName)
                    continue
                else:
                    tStudent = tStudentObj[0]
                # print(tStudent.student_name)
                tTeacherName = sheet_obj.cell(row = i, column = 5).value
                
                if "扶文达f1-6" == tStudentName:
                    tStudentName = "扶文达"
                tTeacherObj = Teacher.objects.filter(display_name = tTeacherName)
                if not tTeacherObj:
                    print(i, tTeacherName)
                    continue
                else:
                    tTeacher = tTeacherObj[0]
                tItemTxt = sheet_obj.cell(row = i, column = 3).value
                tIsPraiseStr = sheet_obj.cell(row = i, column = 2).value
                tIsPraise = 1 if "表扬" == tIsPraiseStr else 0
                tScore = sheet_obj.cell(row = i, column = 4).value
                tDate = sheet_obj.cell(row = i, column = 6).value
                # print(i, tStudentName, tDate)
                oldEvaluateItem = EvaluateItem.objects.filter(item_txt = tItemTxt, is_praise = tIsPraise, score = tScore, sclasses_id = tClass.id)
                if oldEvaluateItem:
                    tEvaluateItemsId = oldEvaluateItem[0].id
                else:
                    tEvaluateItem = EvaluateItem(item_txt = tItemTxt, is_praise = tIsPraise, score = tScore, sclasses_id = tClass.id, evaluate_indexs_id = 1)
                    tEvaluateItem.save()
                    tEvaluateItemsId = tEvaluateItem.id

                tEvaluateLog = EvaluateLog(teachers_id = tTeacher.id, students_id = tStudent.id, evaluate_date = tDate, evaluate_items_id = tEvaluateItemsId)
                tEvaluateLog.save()

    dataset = []
    # for i in range(1, max_col + 1):
    #     cell_obj = sheet_obj.cell(row = 3, column = i)
    #     columns.append(cell_obj.value)
        
    # for i in range(4, max_row + 1):
    #     tCol = []
    #     for j in range(1, max_col + 1):
    #         cell_obj = sheet_obj.cell(row = i, column = j)
    #         tCol.append(cell_obj.value)
    #     dataset.append(tCol)

    # sheet_ranges = wb['range names']
    # print(sheet_ranges['D18'].value)
    # sclass_list = Sclass.objects.all()
    context = {"dataset": dataset}
    return render(request, 'picks/index.html', context)

def readStudentData():
    student_wb = load_workbook('/Users/ywj/Sites/care/static/student.xlsx')
    sheet_obj = student_wb.active
    max_row = sheet_obj.max_row

    for i in range(2, max_row + 1):
        studentName = sheet_obj.cell(row = i, column = 1).value
        identityNumber = sheet_obj.cell(row = i, column = 2).value
        sclassesId = sheet_obj.cell(row = i, column = 3).value 
        gender = sheet_obj.cell(row = i, column = 4).value
        oldStudent = Student.objects.filter(identity_number = identityNumber)
        if oldStudent:
            continue
        else:
            tStudent = Student(student_name = studentName, identity_number = identityNumber, sclasses_id = sclassesId, gender = gender)
            tStudent.save()

def readTeacherData():
    teacher_wb = load_workbook('/Users/ywj/Sites/care/static/teacher.xlsx')
    sheet_obj = teacher_wb.active
    max_row = sheet_obj.max_row

    for i in range(2, max_row + 1):
        teacherName = sheet_obj.cell(row = i, column = 1).value
        gender = sheet_obj.cell(row = i, column = 2).value
        isHeadTeacher = sheet_obj.cell(row = i, column = 3).value
        subjectsId = sheet_obj.cell(row = i, column = 4).value
        displayName = sheet_obj.cell(row = i, column = 5).value
        oldTeacher = Teacher.objects.filter(teacher_name = teacherName, is_head_teacher = isHeadTeacher, gender = gender, subjects_id = subjectsId, display_name = displayName)
        if oldTeacher:
            continue
        else:
            tTeacher = Teacher(teacher_name = teacherName, is_head_teacher = isHeadTeacher, gender = gender, subjects_id = subjectsId, display_name = displayName)
            tTeacher.save()

def analysis(request):
    evaluateItems = EvaluateItem.objects.values("item_txt").distinct().all()
    # print(evaluateItems)
    tItemStr = ""
    resultStr = ""
    for evaluateItem in evaluateItems:
        tItemStr += " "+evaluateItem["item_txt"]

    # tt = jieba.cut(tItemStr, cut_all=False, HMM=True)
    # print("/ ".join(tt))
    # for word, flag, n3, n4 in tt:
        # resultStr += word + " " + flag+" "+n3 + " " + n4
    # allow_pos = ('v', 'vn', 'vd')
    # for key in analyse.extract_tags(tItemStr, 500, withWeight=True, allowPOS=allow_pos):
    # for key in analyse.extract_tags(tItemStr, 500, withWeight=False):
    # 使用jieba.analyse.extract_tags()参数提取关键字,默认参数为50
        # print key.encode('utf-8')
        # resultStr += str(key)
        # 设置输出编码为utf-8不然在因为win下控制台默认中文字符集为gbk,所以会出现乱码
        # 当withWeight=True时,将会返回number类型的一个权重值(TF-IDF)

    # jieba.enable_paddle()# 启动paddle模式。 0.40版之后开始支持，早期版本不支持
    # tags = jieba.analyse.extract_tags(tItemStr, topK=100)
    # tags = jieba.cut_for_search(tItemStr)  # 搜索引擎模式
    # tItemStr = ",".join(tags)
    context = {"dataset": tItemStr}
    return render(request, 'picks/analysis.html', context)